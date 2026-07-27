"""Builds saas_analysis.xlsx — live-formula model for the Meridian Mobility case."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment

wb = openpyxl.Workbook()

# ---- styles ----
HEAD = Font(name="Arial", bold=True, size=12, color="1A2B3C")
SUB  = Font(name="Arial", bold=True, size=10, color="FFFFFF")
BODY = Font(name="Arial", size=10)
BOLD = Font(name="Arial", size=10, bold=True)
BLUE = Font(name="Arial", size=10, color="0000FF")   # inputs
NOTE = Font(name="Arial", size=9, italic=True, color="6A7B8B")
hdr_fill = PatternFill("solid", fgColor="1A2B3C")
key_fill = PatternFill("solid", fgColor="FFF3B0")
tot_fill = PatternFill("solid", fgColor="EDF1F6")
thin = Side(style="thin", color="D0D7DE")
box = Border(left=thin, right=thin, top=thin, bottom=thin)
CUR = '€#,##0;(€#,##0);-'
PCT = '0.0%'

def style_header_row(ws, row, cols):
    for c in cols:
        cell = ws[f"{c}{row}"]
        cell.font = SUB; cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = box

# ======================================================================
# SHEET 1: Assumptions
# ======================================================================
a = wb.active
a.title = "Assumptions"
a["A1"] = "Meridian Mobility — SaaS Consolidation Model"; a["A1"].font = Font(name="Arial", bold=True, size=14, color="1A2B3C")
a["A2"] = "All current-state figures from the case brief. Blue cells are levers — change them and every sheet recalculates."
a["A2"].font = NOTE
a["A4"] = "Negotiation & scenario levers"; a["A4"].font = HEAD
rows = [
    ("Asana negotiated rate (€/seat/yr)", 1000, "input", "List is €1,200/seat. Base case assumes ~17% off on a 3-yr consolidated deal."),
    ("Asana list rate (€/seat/yr)", 1200, "calc", "Derived: current Asana cost / purchased seats."),
    ("Users migrated off Monday.com + Trello", 350, "calc", "Monday.com active 230 + Trello active 120."),
    ("Of those, already hold Asana (overlap)", 150, "input", "SOFT ESTIMATE. Needs real licence-to-headcount mapping. See notes."),
    ("Jira right-sizing at renewal (%)", 0.0, "input", "Base case: 0%. Brief gives no Jira active count, so no cut is modelled."),
]
r = 5
a["A5"]="Lever"; a["B5"]="Value"; a["C5"]="Note"
style_header_row(a, 5, ["A","B","C"])
for label, val, kind, note in rows:
    r += 1
    a[f"A{r}"] = label; a[f"A{r}"].font = BODY; a[f"A{r}"].border = box
    cell = a[f"B{r}"]; cell.value = val; cell.border = box
    if kind == "input":
        cell.font = BLUE; cell.fill = key_fill
    else:
        cell.font = BODY
    cell.number_format = PCT if "%" in label else '#,##0'
    a[f"C{r}"] = note; a[f"C{r}"].font = NOTE; a[f"C{r}"].border = box
# named cells
ASANA_RATE = "Assumptions!$B$6"
OVERLAP    = "Assumptions!$B$9"
JIRA_CUT   = "Assumptions!$B$10"
a.column_dimensions["A"].width = 40
a.column_dimensions["B"].width = 12
a.column_dimensions["C"].width = 66

# ======================================================================
# SHEET 2: Current state
# ======================================================================
c = wb.create_sheet("Current State")
c["A1"] = "Current state — four PM platforms"; c["A1"].font = HEAD
headers = ["Vendor","Purchased seats","Active users","Annual cost","List €/seat","€/active user","Idle seats","Utilisation"]
for i,h in enumerate(headers):
    col = chr(65+i); c[f"{col}2"] = h
style_header_row(c, 2, [chr(65+i) for i in range(len(headers))])
data = [
    ("Jira", 900, None, 1_050_000),
    ("Asana", 650, 510, 780_000),
    ("Monday.com", 350, 230, 420_000),
    ("Trello", 250, 120, 120_000),
]
r = 2
for name, purch, act, cost in data:
    r += 1
    c[f"A{r}"] = name; c[f"A{r}"].font = BOLD
    c[f"B{r}"] = purch; c[f"B{r}"].font = BLUE
    c[f"C{r}"] = act; c[f"C{r}"].font = BLUE
    c[f"D{r}"] = cost; c[f"D{r}"].font = BLUE; c[f"D{r}"].number_format = CUR
    c[f"E{r}"] = f"=D{r}/B{r}"; c[f"E{r}"].number_format = CUR
    c[f"F{r}"] = f'=IF(C{r}="","n/a",D{r}/C{r})'; c[f"F{r}"].number_format = CUR
    c[f"G{r}"] = f'=IF(C{r}="","n/a",B{r}-C{r})'
    c[f"H{r}"] = f'=IF(C{r}="","n/a",C{r}/B{r})'; c[f"H{r}"].number_format = PCT
    for col in "ABCDEFGH": c[f"{col}{r}"].border = box
c[f"C{r+0}"]  # noop
# total row
tr = r + 1
c[f"A{tr}"] = "Total"; c[f"A{tr}"].font = BOLD
c[f"B{tr}"] = f"=SUM(B3:B{r})"; c[f"B{tr}"].font = BOLD
c[f"D{tr}"] = f"=SUM(D3:D{r})"; c[f"D{tr}"].font = BOLD; c[f"D{tr}"].number_format = CUR
for col in "ABCDEFGH":
    c[f"{col}{tr}"].fill = tot_fill; c[f"{col}{tr}"].border = box
c[f"A{tr+2}"] = "Microsoft Planner is bundled in the existing Microsoft 365 licence — €0 incremental, out of scope."
c[f"A{tr+2}"].font = NOTE
c["C3"].comment = Comment("Brief gives no active-user count for Jira. Flagged as a limitation.", "Shruti")
widths = {"A":13,"B":15,"C":13,"D":13,"E":12,"F":13,"G":11,"H":12}
for col,w in widths.items(): c.column_dimensions[col].width = w

# ======================================================================
# SHEET 3: End state
# ======================================================================
e = wb.create_sheet("End State")
e["A1"] = "End state — standardise on Asana, retire Monday.com + Trello, keep Jira"; e["A1"].font = HEAD
e["A3"] = "Asana seat build-up"; e["A3"].font = BOLD
build = [
    ("Retained Asana active users", "='Current State'!C4"),
    ("Migrated from Monday.com + Trello", "='Current State'!C5+'Current State'!C6"),
    ("Less: already hold Asana (overlap)", f"=-{OVERLAP}"),
    ("End-state Asana seats", "=B4+B5+B6"),
]
rr = 3
for lab, f in build:
    rr += 1
    e[f"A{rr}"] = lab; e[f"A{rr}"].font = BODY if "End-state" not in lab else BOLD; e[f"A{rr}"].border = box
    e[f"B{rr}"] = f; e[f"B{rr}"].border = box
    if "End-state" in lab: e[f"B{rr}"].font = BOLD; e[f"B{rr}"].fill = tot_fill
SEATS = "'End State'!$B$7"

e["A9"] = "End-state annual cost"; e["A9"].font = BOLD
for i,h in enumerate(["Vendor","Seats","Rate/status","Annual cost"]):
    e[f"{chr(65+i)}10"] = h
style_header_row(e, 10, ["A","B","C","D"])
rows_e = [
    ("Asana", f"={SEATS}", f"@ {ASANA_RATE}/seat", f"={SEATS}*{ASANA_RATE}"),
    ("Jira",  "='Current State'!B3", "held flat, less right-sizing", f"='Current State'!D3*(1-{JIRA_CUT})"),
    ("Monday.com", 0, "retired", 0),
    ("Trello", 0, "retired", 0),
]
rr = 10
for name, seats, status, cost in rows_e:
    rr += 1
    e[f"A{rr}"] = name; e[f"A{rr}"].font = BOLD
    e[f"B{rr}"] = seats
    e[f"C{rr}"] = status; e[f"C{rr}"].font = NOTE
    e[f"D{rr}"] = cost; e[f"D{rr}"].number_format = CUR
    for col in "ABCD": e[f"{col}{rr}"].border = box
etr = rr + 1
e[f"A{etr}"] = "End-state total"; e[f"A{etr}"].font = BOLD
e[f"D{etr}"] = f"=SUM(D11:D{rr})"; e[f"D{etr}"].font = BOLD; e[f"D{etr}"].number_format = CUR
for col in "ABCD": e[f"{col}{etr}"].fill = tot_fill; e[f"{col}{etr}"].border = box
END_TOTAL = f"'End State'!$D${etr}"
for col,w in {"A":16,"B":12,"C":28,"D":14}.items(): e.column_dimensions[col].width = w

# ======================================================================
# SHEET 4: Savings bridge
# ======================================================================
s = wb.create_sheet("Savings Bridge")
s["A1"] = "Savings bridge — current to end state (no double counting)"; s["A1"].font = HEAD
for i,h in enumerate(["Lever","Impact","Running total"]):
    s[f"{chr(65+i)}3"] = h
style_header_row(s, 3, ["A","B","C"])
s["A4"]="Current spend"; s["B4"]="='Current State'!D7"; s["C4"]="=B4"
levers = [
    ("Remove unused Asana licences", "=-('Current State'!B4-'Current State'!C4)*'Current State'!E4"),
    ("Retire Trello", "=-'Current State'!D6"),
    ("Retire Monday.com", "=-'Current State'!D5"),
    ("Migrate net-new users to Asana", f"=('Current State'!C5+'Current State'!C6-{OVERLAP})*{ASANA_RATE}"),
    ("Renegotiate Asana rate (retained seats)", f"=-('Current State'!E4-{ASANA_RATE})*'Current State'!C4"),
]
rr = 4
for lab, f in levers:
    rr += 1
    s[f"A{rr}"] = lab; s[f"A{rr}"].font = BODY
    s[f"B{rr}"] = f; s[f"B{rr}"].number_format = CUR
    s[f"C{rr}"] = f"=C{rr-1}+B{rr}"; s[f"C{rr}"].number_format = CUR
    for col in "ABC": s[f"{col}{rr}"].border = box
er = rr + 1
s[f"A{er}"] = "End-state spend"; s[f"A{er}"].font = BOLD
s[f"C{er}"] = f"=C{rr}"; s[f"C{er}"].font = BOLD; s[f"C{er}"].number_format = CUR
for col in "ABC": s[f"{col}{er}"].fill = tot_fill; s[f"{col}{er}"].border = box
# check + savings
s[f"A{er+2}"] = "Total annual saving"; s[f"A{er+2}"].font = BOLD
s[f"C{er+2}"] = f"='Current State'!D7-C{er}"; s[f"C{er+2}"].font = Font(name="Arial", bold=True, size=11, color="2F6FED"); s[f"C{er+2}"].number_format = CUR
s[f"A{er+3}"] = "As % of current spend"; s[f"A{er+3}"].font = BODY
s[f"C{er+3}"] = f"=C{er+2}/'Current State'!D7"; s[f"C{er+3}"].number_format = PCT
s[f"A{er+5}"] = "Reconciliation check (bridge end = End State total)"; s[f"A{er+5}"].font = NOTE
s[f"C{er+5}"] = f'=IF(ROUND(C{er}-{END_TOTAL},0)=0,"OK","MISMATCH")'; s[f"C{er+5}"].font = NOTE
for col,w in {"A":40,"B":14,"C":15}.items(): s.column_dimensions[col].width = w

# ======================================================================
# SHEET 5: Scenario range
# ======================================================================
sc = wb.create_sheet("Scenario Range")
sc["A1"] = "Savings range — sensitivity to the Asana rate and Jira right-sizing"; sc["A1"].font = HEAD
for i,h in enumerate(["Scenario","Asana rate","Jira cut","End-state spend","Annual saving","% saving"]):
    sc[f"{chr(65+i)}3"] = h
style_header_row(sc, 3, [chr(65+i) for i in range(6)])
scen = [
    ("Conservative", 1100, 0.0, "Weak Asana deal (−8%), Jira untouched. Matches the brief's ~€540K."),
    ("Base", 1000, 0.0, "3-yr consolidated deal (−17%), Jira untouched."),
    ("Optimistic", 950, 0.10, "Strong deal (−21%) + Jira right-sized 10% at its renewal."),
]
rr = 3
for name, rate, cut, note in scen:
    rr += 1
    sc[f"A{rr}"] = name; sc[f"A{rr}"].font = BOLD
    sc[f"B{rr}"] = rate; sc[f"B{rr}"].font = BLUE; sc[f"B{rr}"].fill = key_fill; sc[f"B{rr}"].number_format = CUR
    sc[f"C{rr}"] = cut; sc[f"C{rr}"].font = BLUE; sc[f"C{rr}"].fill = key_fill; sc[f"C{rr}"].number_format = PCT
    sc[f"D{rr}"] = f"={SEATS}*B{rr}+'Current State'!D3*(1-C{rr})"; sc[f"D{rr}"].number_format = CUR
    sc[f"E{rr}"] = f"='Current State'!D7-D{rr}"; sc[f"E{rr}"].number_format = CUR
    sc[f"F{rr}"] = f"=E{rr}/'Current State'!D7"; sc[f"F{rr}"].number_format = PCT
    for col in "ABCDEF": sc[f"{col}{rr}"].border = box
sc[f"A{rr+2}"] = "Range headline: €540K (conservative) to €750K (optimistic); base case €610K."
sc[f"A{rr+2}"].font = NOTE
for col,w in {"A":15,"B":12,"C":10,"D":16,"E":14,"F":10}.items(): sc.column_dimensions[col].width = w

wb.save("saas_analysis.xlsx")
print("saved saas_analysis.xlsx")
